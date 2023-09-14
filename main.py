# coding=utf-8
import openpyxl
import requests
import wordcloud
import jieba
import re
from bs4 import BeautifulSoup

# 获取Top20弹幕词云
def get_wordcloud(file):
    w = wordcloud.WordCloud(width=1000, height=700, background_color='white', font_path='msyh.ttc')

    f = open(file, encoding='gbk')
    barrage = f.read()
    barrage_list = jieba.lcut(barrage)
    barrage_string = ' '.join(barrage_list)
    w.generate(barrage_string)
    w.to_file('output.png')
    f.close()

# 输出当前弹幕字典中弹幕数量最多的弹幕及其数量
def find_max_num(dict):
    max_num = max(dict.values())
    for name in dict.keys():
        if dict[name] == max_num:
            file_write(name, max_num)
            print(f"{name}:{max_num}")
            dict[name] = 0
            break
    return dict


# 将特定弹幕写入TXT文档中特定次数，用于可视化
def file_write(word, num):
    with open('barrages.txt', 'a') as f:
        for i in range(num):
            f.write(word)
            f.write('\n')


# Excel标题行写入函数
def write_line_excel(arrli):
    work_book = openpyxl.Workbook()
    sheet = work_book.create_sheet('output')

    for i in range(len(arrli)):
        sheet.cell(1, i + 1, arrli[i])

    work_book.save('output.xlsx')


# 将各个弹幕及其数量写入Excel表格
def write_excel(dict):
    work_book = openpyxl.load_workbook('output.xlsx')
    sheet = work_book['output']
    work_row = 2

    for key, value in dict.items():
        sheet.cell(work_row, 1, key)
        sheet.cell(work_row, 2, value)
        work_row += 1

    work_book.save('output.xlsx')


# 获取当前网页视频的所有bvid
def get_bvids(html):
    bvids = []
    response = re.findall(r'bvid:"([^"]+)"', html)
    for bvid in response:
        bvids.append(bvid)
    return bvids


# 通过视频的bvid获取视频的cid（形参为bvid列表，返回值为cid字典）
def get_cid(bvids):
    base_url = "https://api.bilibili.com/x/player/pagelist"

    cids = {}

    for video_id in bvids:
        params = {
            "bvid": video_id
        }

        response = requests.get(base_url, params=params)
        data = response.json()

        if "data" in data and len(data["data"]) > 0:
            cid = data["data"][0]["cid"]
            cids[video_id] = cid

    return cids


def main():
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36 Edg/116.0.1938.69"
    }

    base_url = "https://search.bilibili.com/all?vt=51592400&keyword=%E6%97%A5%E6%9C%AC%E6%A0%B8%E6%B1%A1%E6%9F%93%E6%B0%B4%E6%8E%92%E6%B5%B7&page="

    # 存放所有300个视频的cid
    cids = []

    for i in range(1, 11):
        new_url = base_url + "i"
        result = requests.get(new_url, headers=headers)
        html = result.text
        html_bvids = get_bvids(html)
        html_cids = get_cid(html_bvids)
        for cid in html_cids.values():
            cids.append(cid)

    # 存放各个弹幕数量的字典
    barrage_num = {}

    # 循环遍历并保存各个视频的弹幕
    for cid in cids:

        # 获取当前视频网站的url
        url = f"https://comment.bilibili.com/{cid}.xml"

        response = requests.get(url, headers=headers)
        response.encoding = 'utf-8'
        xml = response.text

        # 获得当前视频的弹幕
        soup = BeautifulSoup(xml, "xml")
        all_barrages = soup.findAll("d")

        # 计算各个弹幕出现的次数并保存到字典中
        for barrage in all_barrages:
            if barrage.string not in barrage_num.keys():
                barrage_num[barrage.string] = 1
            else:
                barrage_num[barrage.string] += 1

    # 写入Excel标题行
    write_line_excel(['barrage', 'num'])
    # 追加写入弹幕数量
    write_excel(barrage_num)

    # 循环遍历并输出出现次数Top20的弹幕
    for i in range(20):
        barrage_num = find_max_num(barrage_num)

    #获取弹幕词云
    get_wordcloud('barrages.txt')


if __name__ == '__main__':
    main()
